
# -*- coding: cp1252 -*-
"""M�dulo para facilidades no manuseio de recursos comuns no desenvolvimento. Vers�o 5.1"""


def cls():
    import os

    os.system('cls')


def logar(
    mensagem,
    nivel,
    arquivo=None,
    modo=None,
    encoding=None,
    formatacao=None,
    handlers=None,
):
    """Formata e retorna uma string como log.
    Ser� exibido sempre o n�vel em primeira posi��o."""
    # importa recursos do m�dulo logging
    from logging import (
        CRITICAL,
        DEBUG,
        ERROR,
        INFO,
        WARNING,
        basicConfig,
        critical,
        debug,
        error,
        info,
        warning,
    )

    # define um n�vel de log
    nivel = nivel.upper()

    # define configura��es b�sicas de log
    basicConfig(
        level=nivel,
        filename=arquivo,
        filemode=modo,
        encoding=encoding,
        format=formatacao,
        handlers=handlers,
    )

    # executa comando de logging conforme o n�vel:
    if nivel == 'DEBUG':
        debug(mensagem)
    elif nivel == 'INFO':
        info(mensagem)
    elif nivel == 'WARNING':
        warning(mensagem)
    elif nivel == 'ERROR':
        error(mensagem)
    elif nivel == 'CRITICAL':
        critical(mensagem)
    # caso o n�vel n�o corresponder aos n�veis padr�es de logging
    else:
        # retorna mensagem de par�metro inv�lido
        return 'Par�metro n�vel inv�lido. Por favor, informe-o corretamente.'

    # retorna a mensagem e o n�vel
    return (nivel, mensagem)


def criar_pasta(caminho):
    """Cria pasta com caminho e nome informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # trata o caminho com o objeto Path
    caminho_interno = Path(caminho)

    # cria a pasta informada, caso necess�rio cria
    #   a hierarquia anterior � �ltima pasta
    caminho_interno.mkdir(parents=True)

    # retorna True caso a opera��o tenha conclu�da com sucesso
    return True


def excluir_pasta(caminho, vazia: bool = True):
    """Exclui pasta no caminho informado. Caso a pasta n�o esteja vazia,
    informe explicitamente no par�metro 'vazia'."""
    # Se a pasta estiver vazia
    if vazia == True:
        # importa recursos do m�dulo Path
        from pathlib import Path

        # trata o caminho com o objeto Path
        caminho_interno = Path(caminho)
        # caso o caminho existir
        if caminho_interno.exists():
            # exclui a pasta informada
            caminho_interno.rmdir()
    # Se a pasta n�o estiver vazia
    elif vazia == False:
        # importa recursos do m�dulo rmtree
        from shutil import rmtree

        # exclui a pasta informada e o conte�do contido nela
        rmtree(caminho)

    # retorna True caso a opera��o tenha conclu�da com sucesso
    return True


def excluir_arquivo(caminho):
    """Exclui um arquivo no caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # trata o caminho com o objeto Path
    arquivo = Path(caminho)

    # se o arquivo existir
    if arquivo.exists():
        # exclui o arquivo informado
        arquivo.unlink()

    # retorna True caso a opera��o tenha conclu�da com sucesso
    return True


def pasta_existente(caminho):
    """Verifica se uma pasta no caminho informado existe."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # verifica e retorna se a pasta existe.
    #   True caso exista e False se n�o existir.
    return Path(caminho).exists()


def pasta_esta_vazia(caminho):
    """Verifica se uma pasta no caminho informado est� vazia."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # coleta de forma recursiva o conte�do
    #   contido no caminho informado caso existir
    lista_arquivos_pastas = list(Path(caminho).glob('**/*'))

    # se n�o existir conte�do no caminho informado
    if len(lista_arquivos_pastas) == 0:
        # retorna True informando que a pasta est� vazia
        return True
    # se existir conte�do no caminho informado
    else:
        # retorna False informando que a pasta n�o est� vazia
        return False


def arquivo_existente(caminho):
    """Verifica se um arquivo no caminho informado existe."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # coleta o caminho absoluto do caminho informado
    arquivo = Path(caminho).absolute()

    # verifica e retorna se o arquivo existe.
    #   True caso exista e False se n�o existir.
    return arquivo.exists()


def abrir_arquivo_texto(caminho, encoding='utf8'):
    """Abre um arquivo de texto no caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # abre um arquivo de texto e coleta o conte�do
    arquivo = Path(caminho).read_text(encoding=encoding)

    # retorna o conte�do do arquivo
    return arquivo


def coletar_nome_guias_arquivo_excel(arquivo_excel):
    """Coleta as guias existentes no arquivo Excel informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # importa recursos do m�dulo openpyxl
    from openpyxl import load_workbook

    # trata o caminho com o objeto Path
    caminho_excel = Path(arquivo_excel).absolute()

    # abre um arquivo de Excel e coleta o conte�do
    conteudo_excel = load_workbook(caminho_excel)

    # define um valor padr�o e inicial � lista
    lista_guias = []

    # coleta a lista de guias que o arquivo cont�m
    lista_guias = conteudo_excel.sheetnames

    # retorna a lista coletada
    return lista_guias


def abrir_arquivo_excel(
    arquivo_excel :  str,
    guia : str = '',
    manter_macro : bool = True,
    manter_links : bool = True,
):
    """Abre um arquivo de Excel no caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # importa recursos do m�dulo openpyxl
    from openpyxl import load_workbook

    # trata o caminho com o objeto Path
    caminho_excel = Path(arquivo_excel).absolute()

    # abre um arquivo de Excel e coleta o conte�do
    conteudo_excel = load_workbook(
        caminho_excel,
        keep_vba=manter_macro,
        keep_links=manter_links
    )

    # seleciona a guia � trabalhar
    if guia == '':
        aba_ativa = conteudo_excel.active
    else:
        aba_ativa = conteudo_excel[guia]

    # define um valor padr�o e inicial � lista
    tabela_excel = []

    # para cada linha do conte�do coletado
    for linhas in aba_ativa.values:
        
        # define um valor padr�o e inicial � lista
        linha = []
        
        # para cada valor na c�lula da linha
        for celula in linhas:
            # adiciona o valor na linha
            linha.append(celula)

        # adiciona a linha na tabela
        tabela_excel.append(linha)

    # retorna o conte�do da tabela
    return tabela_excel


def abrir_arquivo_word(arquivo_word :  str,):
    """Abre um arquivo de word no caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # importa recursos do m�dulo docx
    from docx import Document

    # trata o caminho com o objeto Path
    caminho_word = Path(arquivo_word).absolute()

    # abre um arquivo de word e coleta o conte�do
    conteudo_word = Document(caminho_word)

    # separa o conte�do por par�grafos
    paragrafos = conteudo_word.paragraphs

    # define um valor padr�o e inicial �s listas
    lista_texto_paragrafos = [['texto dos par�grafos']]
    lista_propriedades_paragrafos = [
            [
                'nome do estilo',
                'id do estilo',
                'nome da fonte',
                'tamanho da fonte',
            ]
        ]
    tabela_word = [
        ['texto dos par�grafos'],
        ['propriedades dos par�grafos'],
    ]

    # para cada par�grafo do conte�do coletado
    for paragrafo in paragrafos:

        # salva o texto do par�grafo em quest�o em uma lista
        lista_texto_paragrafos.append(paragrafo.text)

        # salva as propriedades de estiliza��o
        #   do par�grafo em quest�o em uma lista
        lista_propriedades_paragrafos.append(
            [
                paragrafo.style.name,
                paragrafo.style.style_id,
                paragrafo.style.font.name,
                paragrafo.style.font.size,
            ]
        )

    # adiciona as listas na tabela
    tabela_word[0].append(lista_texto_paragrafos)
    tabela_word[1].append(lista_propriedades_paragrafos)

    # retorna o conte�do da tabela
    return tabela_word


def abrir_arquivo_pdf(
    arquivo_pdf :  str,
    senha_pdf: str or None = None,
    paginacao: int or tuple[int] = 0,
    orientacao: int or tuple[int] = (0, 90, 180, 270)
):
    """Abre um arquivo de word no caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # importa recursos do m�dulo PyPDF2
    from PyPDF2 import PdfReader

    # trata o caminho com o objeto Path
    caminho_pdf = Path(arquivo_pdf).absolute()
    # abre um arquivo de PDF e coleta o conte�do
    conteudo_pdf = PdfReader(
        stream=caminho_pdf,
        password=senha_pdf,
        strict=False
    )

    # define um valor padr�o e inicial � lista
    lista_paginacao = []

    # caso o tipo do par�metro 'paginacao' for igual � int
    if type(paginacao) == type(int()):
        # transforma 'paginacao' em tupla
        paginacao = (paginacao,)

    # para cada valor do �ndice de 'pagina��o'
    for indice in paginacao:
        # caso �ndice seja do tipo int
        if type(indice) == type(int()):
            # se �ndice for igual � 0
            if indice == 0:
                # adiciona todas as p�ginas para a lista de pagina��es
                lista_paginacao = conteudo_pdf.pages

                # encerra a itera��o
                break
            # se �ndice n�o for igual � 0
            else:
                # ajusta o �ndice ao padr�o de �ndice de listas
                indice = indice - 1

                # adiciona a pagina��o solicitada � lista de pagina��es
                lista_paginacao.append(conteudo_pdf.getPage(indice))
        # caso �ndice n�o seja do tipo int
        else:
            # levanta exce��o de tipo incorreto
            raise TypeError(
                'Par�metro ``paginacao`` aceita somente �tens num�ricos (int)'
            )

    # define um valor padr�o e inicial � lista
    lista_texto_pdf = []

    # para cada p�gina do conte�do coletado
    for pagina in lista_paginacao:
        # adiciona o valor na linha separando por p�ginas [n] e linhas [n][n]
        lista_texto_pdf.append(
            pagina.extract_text(orientations=orientacao).split('\n')
        )

    # retorna o conte�do coletado em lista
    return lista_texto_pdf


def converter_pdf_em_imagem(
    arquivo_pdf : str,
    caminho_saida: str,
    alpha: bool = False,
    zoom: float = 1,
    orientacao: int or tuple[int] = (0, 90, 180, 270)[0],
):
    """Converte cada p�gina de um arquivo PDF em imagem."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # importa recursos do m�dulo fitz
    import fitz

    try:
        # trata os caminhos com o objeto Path
        caminho_pdf = Path(arquivo_pdf).absolute()
        caminho_saida_img = Path(caminho_saida).absolute()

        # abre um arquivo de PDF e coleta o conte�do
        conteudo_pdf = fitz.open(caminho_pdf)

        # para cada p�gina
        for indice in range(conteudo_pdf.page_count):
            # coleta a p�gina atual
            pagina = conteudo_pdf[indice]
            # coleta a rotacao
            rotacao = orientacao
            # coleta o zoom do eixo X e eixo Y
            zoom_x = zoom_y = zoom
            # trata o arquivo de sa�da
            arquivo_img = caminho_saida_img / f'arquivo_{str(indice+1)}.png'

            # coleta a matriz da p�gina, combinando o zoom e a rota��o
            matriciado = fitz.Matrix(zoom_x, zoom_y).prerotate(rotacao)
            # converte a matriz da p�gina em um mapa
            #   de p�xel de imagem adicionando fundo
            mapa_pixel = pagina.get_pixmap(matrix=matriciado, alpha=alpha)

            # salva o mapa de p�xel em um arquivo de imagem
            mapa_pixel.save(arquivo_img)
        # retorna True caso a opera��o tenha sucesso
        return True
    except:
        # retorna False caso a opera��o tenha sucesso
        return False
        ...


def extrair_texto_ocr(arquivo, linguagem, encoding='utf8'):
    """Extrai texto de arquivo de imagem usando OCR."""
    # importa recursos do m�dulo subprocess
    import subprocess

    # importa recursos do m�dulo Path
    from pathlib import Path

    # abre um arquivo de texto e coleta o conte�do em bytes
    caminho_arquivo = Path(arquivo).absolute()

    # coleta o texto da imagem usando Pytesseract OCR
    texto_extraido = subprocess.run(
        ('pytesseract', '-l', linguagem, caminho_arquivo),
        stdout=subprocess.PIPE,
        encoding=encoding,
    )

    # retorna o texto coletado
    return texto_extraido.stdout


def abrir_arquivo_em_bytes(caminho):
    """Abre em bytes um arquivo de texto no caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # abre um arquivo de texto e coleta o conte�do em bytes
    arquivo = Path(caminho).read_bytes()

    # retorna o conte�do do arquivo
    return arquivo


def criar_arquivo_texto(
    caminho,
    dado='',
    encoding='utf8',
    em_bytes : bool = False,
):
    # importa recursos do m�dulo Path
    """Cria um arquivo de texto no caminho informado."""
    from pathlib import Path

    # caso em_bytes n�o for verdadeiro
    if em_bytes is False:
        # escreve em um arquivo de texto o conte�do informado
        Path(caminho).write_text(encoding=encoding, data=dado)
    # caso em_bytes for verdadeiro
    else:
        # escreve em um arquivo de texto o conte�do informado em bytes
        Path(caminho).write_bytes(data=dado)

    # retorna True caso a opera��o tenha conclu�da com sucesso
    return True


def gravar_log_em_arquivo(
    arquivo,
    conteudo,
    modo,
    encoding='utf8',
    delimitador = ';',
    nova_linha = False,
):
    """salva o conte�do informado em um arquivo de texto tamb�m informado."""
    # transforma todos os argumentos em lista
    lista = list(conteudo)
    # inverte os valores na lista
    lista.reverse()

    # define a variavel
    lista_montada = ''
    # enquanto a lista contiver algum conte�do
    while len(lista) > 0:
        # coleta o �ltimo conte�do da lista
        item = lista.pop()
        # se a lista estiver vazia
        if len(lista) == 0:
            # salva o item coletado na lista instanciada
            lista_montada += item
        # se a lista ainda n�o estiver vazia
        else:
            # salva o item coletado na lista
            #   instanciada e adiciona o delimitador ao final
            lista_montada += item + delimitador

    log = open(arquivo, modo, encoding=encoding)
    if nova_linha == True:
        log.write(lista_montada + '\r\n')
    else:
        log.write(lista_montada + '\r')
    log.close()


def coletar_nome_arquivo(caminho):
    """Coleta o nome de um arquivo no caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # se o arquivo existir
    if Path(caminho).exists() == True:
        # coleta o nome do arquivo informado
        arquivo = Path(caminho).stem
    # caso o arquivo n�o existir
    else:
        # retorna False
        return False

    # retorna o nome do arquivo
    return arquivo


def coletar_extensao_arquivo(caminho):
    """Coleta a extens�o de um arquivo no caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # caso o caminho existir
    if Path(caminho).exists() == True:
        # coleta a extens�o do arquivo
        arquivo = Path(caminho).suffixes
    # caso n�o exista arquivo
    else:
        # retorna False
        return False

    # retorna a extens�o coletada
    return arquivo


def retornar_arquivos_em_pasta(caminho, filtro='**/*'):
    """Retorna os arquivos existentes em um diret�rio se seus
    respectivos sub-diret�rios segundo o caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # coleta de forma recursiva o conte�do
    #   contido no caminho informado caso existir
    lista_arquivos = list(Path(caminho).glob(filtro))

    # instancia uma lista vazia
    lista_arquivos_str = []
    
    # para cada arquivo na lista de arquivos
    for arquivo in lista_arquivos:
        # coleta e salva o arquivo em string
        lista_arquivos_str.append(str(arquivo))

    # retorna uma lista dos arquivos coletados
    return lista_arquivos_str


def coletar_caminho_absoluto(caminho):
    """Retorna os arquivos existentes em um diret�rio se seus
    respectivos sub-diret�rios segundo o caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # coleta o caminho informado no padr�o do objeto Path
    caminho_interno = Path(caminho)
    
    # coleta o caminho absoluto do caminho
    caminho_absoluto = str(caminho_interno.absolute())
    
    # retorna o caminho absoluto coletado
    return caminho_absoluto


def coletar_arvore_caminho(caminho):
    """Retorna os arquivos existentes em um diret�rio se seus
    respectivos sub-diret�rios segundo o caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # coleta o caminho informado no padr�o do objeto Path
    caminho_interno = Path(caminho)
    
    # coleta a �rvore do caminho informado
    arvore_caminho = str(caminho_interno.parent)
    
    # retorna o caminho absoluto coletado
    return arvore_caminho


def renomear(caminho, nome_atual, novo_nome):
    """Renomeia o nome de um arquivo no caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # trata o caminho informado e o nome atual com o objeto Path
    nome_atual = Path(caminho) / nome_atual

    # trata o caminho informado e o nome novo com o objeto Path
    novo_nome = Path(caminho) / novo_nome

    # valida se o nome atual existe
    validacao_nome_atual = nome_atual.exists()

    # valida se o nome atual existe
    validacao_novo_nome = novo_nome.exists()

    # se o nome novo n�o existir e o nome atual existir
    if not (validacao_novo_nome == True) and (validacao_nome_atual == True):
        # altera o nome atual para o nome novo
        novo_nome = nome_atual.rename(novo_nome)

        # retorna o objeto Path com o nome novo
        return novo_nome
    # se o nome novo existir ou o nome atual n�o existir
    else:
        # retorna False
        return False


def recortar(caminho_atual, caminho_novo):
    """Recorta um arquivo ou pasta de um caminho
    e cola em outro caminho conforme informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # trata o caminho atual com o objeto Path
    caminho_atual = Path(caminho_atual)
    # trata o caminho novo com o objeto Path
    caminho_novo = Path(caminho_novo)

    # valida se o caminho atual existe
    validacao_caminho_atual = caminho_atual.exists()

    # valida se o caminho novo existe
    validacao_caminho_novo = caminho_novo.exists()
    # se o caminho atual existir e o caminho novo n�o existir
    if (validacao_caminho_atual == True) and not (
        validacao_caminho_novo == True
    ):
        # modifica o nome conforme informado
        caminho_novo = caminho_atual.rename(caminho_novo)
        # retorna o objeto Path com o caminho novo
        return caminho_novo
    # caso o caminho atual n�o exista ou o caminho novo j� existir
    else:
        # retorna False
        return False


def copiar_arquivo(arquivo, caminho_destino):
    """Copia um arquivo de um caminho para
    outro caminho conforme informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # trata o arquivo com o objeto Path
    arquivo = Path(arquivo)

    # se o arquivo existir
    if arquivo.exists() == True:
        # coleta o caminho absoluto do arquivo
        arquivo = arquivo.absolute()
    # se o arquivo n�o existir
    else:
        # retorna False
        return False

    # trata o caminho de destino com o objeto Path
    caminho_destino = Path(caminho_destino)
    # se o caminho existir
    if caminho_destino.exists() == True:
        # importa recursos do m�dulo shutil
        from shutil import copy2

        # copia o arquivo para a pasta de destino informado
        caminho_destino = copy2(arquivo, caminho_destino)

        # retorna o caminho de destino
        return caminho_destino
    # se o caminho n�o existir
    else:
        # retorna False
        return False


def copiar_pasta(pasta, caminho_destino):
    """Copia uma pasta de um caminho para outro caminho conforme informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path

    # trata o caminho com o objeto Path
    pasta_var_interna = Path(pasta)
    # se a pasta existir
    if pasta_var_interna.exists() == True:
        # trata o caminho de destino com o objeto Path
        caminho_destino_var_interna = Path(caminho_destino)
        # se o caminho de destino existir
        if caminho_destino_var_interna.exists() == True:
            # importa recursos do m�dulo shutil
            from shutil import copytree

            # copia a pasta para o destino informado
            caminho_destino = copytree(
                pasta, caminho_destino_var_interna / pasta
            )
        # se o caminho de destino n�o existir
        else:
            # retorna False
            return False
    # se a pasta n�o existir
    else:
        # retorna False
        return False

    # retorna o caminho de destino com a pasta copiada
    return caminho_destino


def descompactar(arquivo, caminho_destino):
    """Descompacta um arquivo para o caminho informado."""
    # importa recursos do m�dulo Path
    from pathlib import Path
    # importa recursos do m�dulo subprocess
    import subprocess


    # trata o arquivo com o objeto Path
    arquivo_interno = Path(arquivo)
    # trata o caminho de destino com o objeto Path
    caminho_interno = Path(caminho_destino)

    # coleta o caminho absoluto do arquivo e trata para string
    arquivo_interno_absoluto = arquivo_interno.absolute()
    arquivo_interno_absoluto_str = str(arquivo_interno_absoluto)

    # coleta o caminho absoluto do arquivo e trata para string
    caminho_interno_absoluto = caminho_interno.absolute()
    caminho_interno_absoluto_str = str(caminho_interno_absoluto)

    try:
        # descompacta o arquivo
        subprocess.call(
            [
                "powershell", f'''Expand-Archive -Path "{
                    arquivo_interno_absoluto_str
                }" -DestinationPath "{
                    caminho_interno_absoluto_str
                }"  | Out-Null'''
            ],
            stdin=False,
            stdout=False,
            stderr=False,
            shell=False,
            cwd=None,
            timeout=None,
            text=False,
        )

        # retorna True
        return True
    except:
        # retorna False
        return False


def ler_variavel_ambiente(
    arquivo_config='config.ini',
    nome_bloco_config='padrao',
    nome_variavel=None,
    variavel_sistema: bool = False,
    encoding='utf8',
):
    """L� uma vari�vel de ambiente,
    tanto de um arquivo quanto direto do sistema."""
    # importa recursos do m�dulo os
    import os
    # importa recursos do m�dulo ConfigParser
    from configparser import ConfigParser

    # se n�o for vari�vel de sistema
    if not variavel_sistema == True:
        # instancia o objeto de configura��o
        config = ConfigParser()
        # L� o arquivo de configura��o
        config.read(arquivo_config, encoding=encoding)

        # se o nome da vari�vel de ambiente foi informada
        if not nome_variavel == None:
            # coleta o dado da vari�vel de ambiente informado
            bloco = dict(config[nome_bloco_config])
            # retorna o valor coletado
            return bloco[nome_variavel]
        # se o nome da vari�vel de ambiente n�o foi informada
        else:
            # retorna o todos os dados no
            #   bloco de vari�veis de ambiente informado
            return dict(config[nome_bloco_config])
    # se for vari�vel de sistema
    else:
        # retorna o valor da vari�vel de sistema solicitado
        return os.environ.get(nome_variavel)


def formatar_log(*args, delimitador=';'):
    """Formata e retorna uma string com formato para log."""
    # transforma todos os argumentos em lista
    lista = list(args)
    # inverte os valores na lista
    lista.reverse()

    # define a variavel
    lista_montada = ''
    # enquanto a lista contiver algum conte�do
    while len(lista) > 0:
        # coleta o �ltimo conte�do da lista
        item = lista.pop()
        # se a lista estiver vazia
        if len(lista) == 0:
            # salva o item coletado na lista instanciada
            lista_montada += item
        # se a lista ainda n�o estiver vazia
        else:
            # salva o item coletado na lista
            #   instanciada e adiciona o delimitador ao final
            lista_montada += item + delimitador

    # returna o n�vel de log adicionando ao final a vari�vel formatada
    return lista_montada


def retornar_data_hora_atual(parametro, separador='/'):
    """Formata e retorna dados de data e/ou hora,
    conforme informado pelo par�metro."""
    # importa recursos do m�dulo datetime
    import datetime

    # retorna dados de data e/ou hora conforme informado pelo par�metro.
    return datetime.datetime.now().strftime(parametro).replace('/', separador)


def coletar_idioma_so():
    """Coleta o idioma atual do sistema operacional."""
    # importa recursos do m�dulo ctypes
    import ctypes

    # importa recursos do m�dulo locale
    import locale

    # coleta as informa��es do kernel do Windows
    windows_dll = ctypes.windll.kernel32

    # coleta o valor do idioma local do sistema no padr�o de ID numérico
    windows_dll.GetUserDefaultUILanguage()

    # coleta o valor do idioma local do sistema no padr�o de escrita abreviada
    idioma = locale.windows_locale[windows_dll.GetUserDefaultUILanguage()]

    # retorna o valor de idioma coletado
    return idioma


def janela_existente(nome_janela):
    """Coleta o idioma atual do sistema operacional."""
    # importa recursos do m�dulo psutil
    import psutil

    # para cada processo na lista de processos
    for processo in psutil.process_iter():
        # tenta executar a a��o
        try:
            # verifica se o nome do processo corresponde ao nome informado
            if nome_janela.lower() in processo.name().lower():
                # caso exista retorna True
                return True
        # para a lista de erros informados
        except (
            psutil.NoSuchProcess,
            psutil.AccessDenied,
            psutil.ZombieProcess,
        ):
            # ignora os erros
            ...
    # retorna False caso n�o encontre processo com o nome informado
    return False


def coletar_pid(nome_processo):
    """Coleta o idioma atual do sistema operacional."""
    # importa recursos do m�dulo os
    import psutil

    # instancia uma lista vazia
    listaProcessos = []
    # para cada processo na lista de processos
    for processo in psutil.process_iter():
        # tenta executar a a��o
        try:
            # coleta o PID, o nome, o tempo de in�cio do processo
            informacao_processo = processo.as_dict(
                attrs=['pid', 'name', 'create_time']
            )
            # se existir um processo com o mesmo nome informado
            if nome_processo.lower() in informacao_processo['name'].lower():
                # salva o nome do processo
                listaProcessos.append(informacao_processo)

        # para a lista de erros informados
        except (
            psutil.NoSuchProcess,
            psutil.AccessDenied,
            psutil.ZombieProcess,
        ):
            # ignora os erros
            ...
    # retorna uma lista de dicion�rios com o nome do processo coletado
    return listaProcessos


def finalizar_processo(pid):
    """Coleta o idioma atual do sistema operacional."""
    # importa recursos do m�dulo os
    import psutil

    # instancia um dicion�rio vazio
    listaProcessos = {}
    # para cada processo na lista de processos
    for processo in psutil.process_iter():
        # tenta executar a a��o
        try:
            # coleta o PID, o nome, o tempo de in�cio do processo
            informacao_processo = processo.as_dict(
                attrs=['pid', 'name', 'create_time']
            )
            # se existir um processo com o mesmo nome informado
            if pid == informacao_processo['pid']:
                # encerra o processo informado
                processo.kill(**informacao_processo)
                # retorna true
                return True
        # para a lista de erros informados
        except (
            psutil.NoSuchProcess,
            psutil.AccessDenied,
            psutil.ZombieProcess,
        ):
            # ignora os erros
            ...
    # retorna um dicion�rio com o nome do processo coletado
    return listaProcessos


def janela_dialogo(titulo, texto, estilo):
    """Exibe uma janela de mensagem na tela."""
    # importa recursos do m�dulo ctypes
    import ctypes

    # cria o objeto de janela conforme os par�mentros informados
    caixa = ctypes.windll.user32.MessageBoxW(0, texto, titulo, estilo)

    # retorna o objeto
    return caixa
